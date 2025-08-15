import pandas as pd 
import polars as pl
from pathlib import Path
import os
import re
import json



### ////////////////////////////////////////////////////////////////////////////////////////
### domain class 

class Reporter:
    def __init__(self, output):
        self.output = output
        self.settings = {
            "field_col" : "Field",
            "pattern_col" : "Pattern"
        }
        self.enc = {
            "r" : "regimen","rc" : "regimenCUI", "vc" : "variantCUI","v" : "variant",'co':"component",
            "c" : "condition","cc" : "conditionCUI", "ad" : "allDays", "lb" : "cycleLengthLB",
             "ub" : "cycleLengthUB", "unit" : "cycleLengthUnit", "ts" : "timingSequence",
             "sn":"stepNumber", 'cr': "componentRole"
            # ... 
        }
        self.decode = lambda x: re.sub(r"\{(.*?)\}", self._replace_match, x)
        self.add = {
            "H" : "HANDLED",     # kept as is
            "P" : "PATCHED",     # kept with a temporary fix
            "N" : "NOT HANDLED", # dropped
        }

        os.makedirs(self.output, exist_ok=True)

    def _replace_match(self, match):
        key = match.group(1)  
        return self.enc.get(key, key)  

    def to_tsv(self, frame, file_name):
        frame.write_csv(f"{self.output}/{file_name}.tsv", separator="\t")
    
    def resolve(self, frame, file_name, pattern=None, field=None, status=None):
        if pattern and field and status:
            frame = frame.with_columns([
                pl.lit(self.decode(pattern) ).alias("Pattern"),
                pl.lit(self.decode(field)   ).alias("Field"),
                pl.lit(self.add[status]     ).alias("Status")
            ])
        else:
            raise TypeError(f"Please specify a value for the param:\nINPUT: pattern={pattern} ; field={field} ; status={status}. ")

        frame.write_csv(f"{self.output}/{file_name}.resolved.tsv", separator="\t")


### ////////////////////////////////////////////////////////////////////////////////////////
### adapter functions

def group_adapt_pl(frame):
    variant_cui_nunique = frame.select(["regimen_cui", "variant_cui"]).unique().height
    regimen_cui_nunique = frame.select("regimen_cui").unique().height

    return frame.with_columns([
        pl.lit(variant_cui_nunique).alias("variant_cui_nunique"),
        pl.lit(regimen_cui_nunique).alias("regimen_cui_nunique")
    ])

def get_sum_pl(frame):
    total = frame.height

    status_stats = (
        frame.group_by("Status")
             .agg(pl.count().alias("c"))
             .with_columns([
                 (pl.col("count") / total * 100).round(2).alias("perc"),
                 pl.col("Status").str.slice(0, 1).alias("Status_initial")
             ])
    )

    for row in status_stats.iter_rows(named=True):
        col_name = f"Status_{row['Status_initial']}_perc"
        frame = frame.with_columns(
            pl.lit(row['perc']).alias(col_name)
        )

    return frame

def group_adapt(df):
    # Compute global unique count of regimen_cui
    regimen_cui_nunique = df["regimen_cui"].nunique()
    df["regimen_cui_nunique"] = regimen_cui_nunique
    df["regimen_cui_nunique_sum"] = regimen_cui_nunique # symmetry

    # Compute per-regimen unique variant count
    per_regimen_variant_counts = (
        df[["regimen_cui", "variant_cui"]]
        .drop_duplicates()
        .groupby("regimen_cui")
        .size()
        .reset_index(name="variant_cui_nunique")
    )

    # Merge per-regimen variant counts back into main df
    df = df.merge(per_regimen_variant_counts, on="regimen_cui", how="left")

    # Compute sum of variant counts across all regimens
    variant_cui_nunique_sum = per_regimen_variant_counts["variant_cui_nunique"].sum()
    df["variant_cui_nunique_sum"] = variant_cui_nunique_sum

    return df

def get_sum(df):
    expected_statuses = ["H", "P", "N"]
    # Step 1: Deduplicate variant-regimen combinations
    dedup = df[["regimen_cui", "variant_cui", "Status"]].drop_duplicates()
    # Step 2: Extract status initial
    dedup["Status_initial"] = dedup["Status"].str[0]

    # Step 3: Count how many variants per status per regimen
    counts = (
        dedup.groupby(["regimen_cui", "Status_initial"])
        .size()
        .unstack(fill_value=0)
        .rename(columns=lambda s: f"Status_{s}_count")
    )

    # Step 4: Compute total variants per regimen_cui
    counts["total"] = counts.sum(axis=1)

    # Step 5: Convert counts to percentages
    for s in expected_statuses:
        col = f"Status_{s}_count"
        perc_col = f"Status_{s}_perc"
        if col not in counts.columns:
            counts[col] = 0
        counts[perc_col] = round((counts[col] / counts["total"]) * 100, 2)

    # Step 6: Join back to original dataframe on regimen_cui
    out = df.merge(
        counts[[f"Status_{s}_perc" for s in expected_statuses]].reset_index(),
        on="regimen_cui",
        how="left"
    )

    return out




### ////////////////////////////////////////////////////////////////////////////////////////
### publisher functions

import pandas as pd
from pathlib import Path

def apply_operations(frame, schema, adapters=[]):
    for adapt in adapters:
        frame = adapt(frame)
    frame.rename(columns=schema, inplace=True)
    return frame[[*schema.values()]]

def load(frame_path):
    return pd.read_csv(frame_path, sep='\t', index_col=False, low_memory=False)


from openpyxl import load_workbook
from pathlib import Path
import pandas as pd

def write_frame_to_excel_sheet(path, frame, sheet_name, headers=True):
    path = Path(path)

    # If file doesn't exist, write with header
    if not path.exists():
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            frame.to_excel(writer, sheet_name=sheet_name, index=False, header=headers)
        return

    # Load existing workbook and determine startrow
    book = load_workbook(path)
    startrow = book[sheet_name].max_row if sheet_name in book.sheetnames else 0

    # Reopen writer without assigning writer.book manually
    with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        # Ensure correct worksheet context
        if sheet_name in writer.sheets:
            writer.sheets[sheet_name] = book[sheet_name]

        frame.to_excel(
            writer,
            sheet_name=sheet_name,
            index=False,
            header=(headers if startrow == 0 else False),
            startrow=(startrow if not headers else 0)
        )


def build_reports(sheets, report_tables_path: str, output_dir: str):
    report_tables_path = Path(report_tables_path)
    output_dir = Path(output_dir)

    output_dir.mkdir(exist_ok=True)
    out_file = output_dir / "reports.xlsx"
    sheets = json.load(open(sheets, "r"))['sheets']
    
    if out_file.exists():# Dev note: default cleanup in case xlsx corrupted 
        os.remove(out_file)

    print("[INFO] Sheets processed:")
    for sheet in sheets:
        print(sheet)
        sheet_name = sheet["name"]
        schema = sheet["headers"]
        adapters = [group_adapt, get_sum]

        # write headers first
        headers_df = pd.DataFrame(columns=[*schema.values()])
        write_frame_to_excel_sheet(out_file, headers_df, sheet_name)

        for table in report_tables_path.glob("*.resolved.tsv"):
            frame = load(table)
            frame = apply_operations(frame, schema, adapters)
            frame = frame[headers_df.columns].drop_duplicates()
            write_frame_to_excel_sheet(out_file, frame, sheet_name, headers=False)
